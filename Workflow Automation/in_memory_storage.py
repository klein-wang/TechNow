import io
import pandas as pd
from openpyxl import Workbook, load_workbook

# 创建一个示例数据框
data = {'Name': ['John', 'Mary', 'David'], 
        'Age': [25, 31, 42]}
df = pd.DataFrame(data)

# 使用 StringIO 存储 CSV 文件
with io.StringIO() as f:
    # 将数据框写入内存文件
    df.to_csv(f, index=False)
    # 获取内存文件的内容
    contents_csv = f.getvalue()

# 从字符串中检索数据
df_retrieved_csv = pd.read_csv(io.StringIO(contents_csv))
print("从 CSV 字符串中检索的数据:")
print(df_retrieved_csv)

# 使用 BytesIO 存储 Excel 文件
with io.BytesIO() as f:
    # 创建一个 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    # 将数据框写入 Excel 工作表
    for r in df.itertuples(index=False):
        ws.append(r)
    # 将工作簿保存到内存文件
    wb.save(f)
    # 获取内存文件的内容
    contents_xlsx = f.getvalue()

# 从二进制流中检索数据
with io.BytesIO(contents_xlsx) as f:
    wb = load_workbook(f)
    ws = wb.active
    df_retrieved_xlsx = pd.DataFrame(ws.values)
print("从 Excel 二进制流中检索的数据:")
print(df_retrieved_xlsx)